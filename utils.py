import pickle
import time
import pandas as pd

import numpy as np
import torch
import torch.utils.data as torch_data
from openpyxl import load_workbook

num2char = {0: "a", 1: "b", 2: "c", 3: "d", 4: "e", 5: "f", 6: "g", 7: "h", 8: "i", 9: "j", 10: "k", 11: "l", 12: "m",
            13: "n", 14: "o", 15: "p", 16: "q", 17: "r", 18: "s", 19: "t", 20: "u"}

char2num = {"a": 0, "b": 1, "c": 2, "d": 3, "e": 4, "f": 5, "g": 6, "h": 7, "i": 8, "j": 9, "k": 10, "l": 11, "m": 12,
            "n": 13, "o": 14, "p": 15, "q": 16, "r": 17, "s": 18, "t": 19, "u": 20}

temperature = 1
Cpuct = 0.1
batch_size = 20

# 棋盘：8*15
# 棋盘大小设为可变，由读入的矩阵大小决定
df = pd.read_excel("0222矩阵.xlsx", header=None)
row_size = df.shape[0]-1  # 属性数量
column_size = df.shape[1]  # 每个属性对应的字段数量
learning_rate = 0.1
cat_size = 5    # 属性数量
max_step = column_size  # 最大步数,因为一列只落一子,所以与column_size相等

class distribution_calculater:
    def __init__(self, row_size, column_size):
        self.map = {}
        self.order = []
        for i in range(row_size):
            for j in range(column_size):
                name = num2char[i] + num2char[j]
                self.order.append(name)
                self.map[name] = 0

    def push(self, key, value):
        self.map[key] = value

    def get(self, train=True):
        result = []
        choice_pool = []
        choice_prob = []
        for key in self.order:
            if self.map[key] != 0:
                choice_pool.append(key)
                tmp = np.float_power(self.map[key], 1 / temperature)
                choice_prob.append(tmp)
                result.append(tmp)
                self.map[key] = 0
            else:
                result.append(0)

        he = sum(result)
        for i in range(len(result)):
            if result[i]:
                result[i] = result[i] / he
        choice_prob = [choice / he for choice in choice_prob]
        if train:
            move = np.random.choice(choice_pool, p=0.8 * np.array(choice_prob) + 0.2 * np.random.dirichlet(
                0.3 * np.ones(len(choice_prob))))
        else:
            move = choice_pool[np.argmax(choice_prob)]

        return move, result

    def getChoiceMoves(self, train=True):
        result = []
        choice_pool = []
        choice_prob = []
        choiceMoves = []
        for key in self.order:
            if self.map[key] != 0:
                choice_pool.append(key)
                tmp = np.float_power(self.map[key], 1 / temperature)
                choice_prob.append(tmp)
                result.append(tmp)
                self.map[key] = 0
            else:
                result.append(0)

        he = sum(result)
        for i in range(len(result)):
            if result[i]:
                result[i] = result[i] / he
        choice_prob = [choice / he for choice in choice_prob]
        if train:
            move = np.random.choice(choice_pool, p=0.8 * np.array(choice_prob) + 0.2 * np.random.dirichlet(
                0.3 * np.ones(len(choice_prob))))
        else:
            move = choice_pool[np.argmax(choice_prob)]
            choice_prob_sortIndex = np.argsort(choice_prob)
            num = len(choice_prob)
            if num > 15:
                num = 15
            for index in range(num):
                oldindex = choice_prob_sortIndex[len(choice_prob) - (index + 1)]
                move = choice_pool[oldindex]
                choiceMove = {}
                choiceMove["action"] = move
                choiceMove["score"] = choice_prob[oldindex]
                choiceMoves.append(choiceMove)

        return choiceMoves, result


def step_child_remove(board_pool, child_pool):
    i = 0
    while i < len(board_pool) and len(child_pool) != 0:
        j = 0
        while j < len(child_pool):
            if np.array_equal(board_pool[i], child_pool[j]):
                board_pool.pop(i)
                child_pool.pop(j)
                i -= 1
                break
            else:
                j += 1
        i += 1
    return board_pool


def write_file(object, file_name):
    filewriter = open(file_name, 'wb')
    pickle.dump(object, filewriter)
    filewriter.close()


def read_file(file_name):
    filereader = open(file_name, 'rb')
    object = pickle.load(filereader)
    filereader.close()
    return object


def move_to_str(action):
    return num2char[action[0]] + num2char[action[1]]


def str_to_move(str):
    return np.array([char2num[str[0]], char2num[str[1]]])


def valid_move(state):
    return list(np.argwhere(state == 0))


def generate_new_state(old_name, step, current_player):
    if current_player == 1:
        step = "B" + num2char[step[0]] + num2char[step[1]]
    else:
        step = "W" + num2char[step[0]] + num2char[step[1]]
    for i in range(0, len(old_name), 3):
        if old_name[i + 1] > step[1] or (old_name[i + 1] == step[1] and old_name[i + 2] > step[2]):
            new_name = old_name[:i] + step + old_name[i:]
            return new_name
    new_name = old_name + step
    return new_name


class random_stack:
    def __init__(self, length=1000):
        self.state = []
        self.distrib = []
        self.winner = []
        self.length = length

    def isEmpty(self):
        return len(self.state) == 0

    def push(self, item):
        self.state.append(item["state"])
        self.distrib.append(item["distribution"])
        self.winner.append(item["value"])
        if len(self.state) >= self.length:
            self.state = self.state[1:]
            self.distrib = self.distrib[1:]
            self.winner = self.winner[1:]

    def seq(self):
        return self.state, self.distrib, self.winner

    # def random_seq(self):
    #     tmp = copy.deepcopy(self.data)
    #     random.shuffle(tmp)
    #     return tmp


# 根据一次对局记录，产生训练数据
def generate_training_data(game_record, row_size, column_size):
    board = np.zeros([row_size, column_size])  # 生成棋盘
    data = []
    player = 1  # 黑方先手，player=1表示黑方，player=-1表示白方
    winner = -1 if len(game_record) % 2 == 0 else 1  # 确定winner是谁，1黑，-1白

    for i in range(len(game_record)):
        step = str_to_move(game_record[i]['action'])  # 获取每一步的落子情况
        state = transfer_to_input(board, player, row_size, column_size)  # 棋局和玩家信息
        data.append({"state": state, "distribution": game_record[i]['distribution'], "value": winner})
        board[step[0], step[1]] = player  # 落子，改变棋盘状态
        player, winner = -player, -winner
    return data


def generate_data_loader(stack):
    state, distrib, winner = stack.seq()
    tensor_x = torch.stack(tuple([torch.from_numpy(s) for s in state]))
    tensor_y1 = torch.stack(tuple([torch.Tensor(y1) for y1 in distrib]))
    tensor_y2 = torch.stack(tuple([torch.Tensor([float(y2)]) for y2 in winner]))
    dataset = torch_data.TensorDataset(tensor_x, tensor_y1, tensor_y2)
    my_loader = torch_data.DataLoader(dataset, batch_size=batch_size, shuffle=True)
    return my_loader


def transfer_to_input(state, current_player, row_size, column_size):
    # 每个tmp是一个矩阵
    if current_player == 1:  # 黑方
        tmp3 = np.ones([row_size, column_size]).astype(float)  # 全1矩阵，表示黑方
        tmp2 = np.array(state > 0).astype(float)  # 对应位置元素>0则为1，否则为0
        tmp1 = np.array(state < 0).astype(float)  # 对应位置元素<0则为1，否则为0
    else:  # 白方
        tmp3 = np.zeros([row_size, column_size])  # 全0矩阵，表示白方
        tmp2 = np.array(state < 0).astype(float)
        tmp1 = np.array(state > 0).astype(float)
    return np.stack([tmp1, tmp2, tmp3])


def visualization(file_name, row_size, column_size):
    action_record = []
    record = read_file(file_name)
    for i in record:
        action_record.append(i["action"])
    board = np.zeros([row_size, column_size])
    stone = 1
    for action in action_record:
        act = str_to_move(action)
        board[act[0], act[1]] = stone
        stone = - stone
        print(board, end="\r")
        time.sleep(2)


def readMatrix(file_name):
    # 打开文件
    wb = load_workbook(file_name)
    sheetNames = wb.sheetnames
    sheet = wb.get_sheet_by_name(sheetNames[0])
    rownum = 0
    head = {}
    disease = []
    matrix = []
    diseaseIndex = 0
    for row in sheet.rows:
        if rownum == 0:
            for index, cell in enumerate(row):
                head[index] = cell.value
                if isDisease(cell.value):
                    diseaseIndex = index
        else:
            column = []
            for index, cell in enumerate(row):
                column.append(cell.value)
                if index == diseaseIndex:
                    disease.append(cell.value)
            matrix.append(column)
        rownum += 1

    return head, disease, matrix


# 考虑到之后会有若干矩阵，每个矩阵中疾病一列不一定叫做疾病，加一个公共方法判断是不是属于类型是不是疾病
def isDisease(str):
    return str == '疾病'


def isContainsDisease(arr):
    return '疾病' in arr


# 计算同一列是否已经有同色的落子, 返回true表示该列可以落子
def can_place_in_this_column(player, matrix, column_index):
    for i in range(len(matrix)):
        if matrix[i][column_index] == player:   # 这一列存在同色落子,则player不可再该列再次落子
            return False
    return True

